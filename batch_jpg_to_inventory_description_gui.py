import re
import threading
from copy import copy
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from PIL import Image, ImageOps, ImageFilter
import pytesseract
from openpyxl import load_workbook


class InventoryOCRApp:
        def _ensure_tesseract(self):
            import shutil
            tesseract = self.tesseract_path.get().strip()
            if tesseract and Path(tesseract).exists():
                pytesseract.pytesseract.tesseract_cmd = tesseract
                return True
            # Try to find tesseract in PATH
            tesseract_in_path = shutil.which("tesseract")
            if tesseract_in_path:
                pytesseract.pytesseract.tesseract_cmd = tesseract_in_path
                return True
            # Prompt user to select tesseract
            messagebox.showwarning(
                "Tesseract Not Found",
                "Tesseract is not installed or not in your PATH. Please select the Tesseract executable."
            )
            self._browse_tesseract()
            tesseract = self.tesseract_path.get().strip()
            if tesseract and Path(tesseract).exists():
                pytesseract.pytesseract.tesseract_cmd = tesseract
                return True
            return False
    def __init__(self, root):
        self.root = root
        self.root.title("JPG OCR to Inventory Template")
        self.root.geometry("920x760")
        self.root.minsize(860, 700)

        self.template_file = tk.StringVar()
        self.output_file = tk.StringVar()
        self.input_mode = tk.StringVar(value="folder")
        self.input_folder = tk.StringVar()
        self.tesseract_path = tk.StringVar()
        self.lang = tk.StringVar(value="eng")
        self.psm = tk.StringVar(value="6")
        self.oem = tk.StringVar(value="3")
        self.sheet_name = tk.StringVar(value="Data")
        self.header_row = tk.StringVar(value="1")
        self.start_row = tk.StringVar(value="2")
        self.target_header = tk.StringVar(value="Description")
        self.clean_line_breaks = tk.BooleanVar(value=False)
        self.add_filename_prefix = tk.BooleanVar(value=False)
        self.insert_filename_item = tk.BooleanVar(value=True)

        self.selected_files = []
        self.is_running = False

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        ttk.Label(
            self.root,
            text="JPG OCR to Inventory Template",
            font=("Arial", 16, "bold")
        ).pack(anchor="w", **pad)

        ttk.Label(
            self.root,
            text="Loads an Excel inventory template and writes one OCR result per image into the Description column."
        ).pack(anchor="w", **pad)

        # Template
        template_frame = ttk.LabelFrame(self.root, text="Excel Template")
        template_frame.pack(fill="x", padx=10, pady=8)

        ttk.Label(template_frame, text="Template workbook").grid(row=0, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(template_frame, textvariable=self.template_file, width=78).grid(
            row=0, column=1, sticky="ew", padx=10, pady=6
        )
        ttk.Button(template_frame, text="Browse", command=self._browse_template).grid(
            row=0, column=2, sticky="ew", padx=10, pady=6
        )

        ttk.Label(template_frame, text="Output workbook").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(template_frame, textvariable=self.output_file, width=78).grid(
            row=1, column=1, sticky="ew", padx=10, pady=6
        )
        ttk.Button(template_frame, text="Save As", command=self._save_output).grid(
            row=1, column=2, sticky="ew", padx=10, pady=6
        )

        template_frame.columnconfigure(1, weight=1)

        # Images
        input_frame = ttk.LabelFrame(self.root, text="Images")
        input_frame.pack(fill="x", padx=10, pady=8)

        ttk.Radiobutton(
            input_frame, text="Use a folder", variable=self.input_mode, value="folder",
            command=self._toggle_input_mode
        ).grid(row=0, column=0, sticky="w", padx=10, pady=6)

        ttk.Radiobutton(
            input_frame, text="Choose individual files", variable=self.input_mode, value="files",
            command=self._toggle_input_mode
        ).grid(row=0, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(input_frame, text="Folder").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        self.folder_entry = ttk.Entry(input_frame, textvariable=self.input_folder, width=78)
        self.folder_entry.grid(row=1, column=1, sticky="ew", padx=10, pady=6)
        self.folder_button = ttk.Button(input_frame, text="Browse Folder", command=self._browse_folder)
        self.folder_button.grid(row=1, column=2, sticky="ew", padx=10, pady=6)

        self.files_button = ttk.Button(input_frame, text="Choose Files", command=self._browse_files)
        self.files_button.grid(row=2, column=2, sticky="ew", padx=10, pady=6)

        self.files_label = ttk.Label(input_frame, text="No files selected")
        self.files_label.grid(row=2, column=1, sticky="w", padx=10, pady=6)

        input_frame.columnconfigure(1, weight=1)

        # Mapping
        map_frame = ttk.LabelFrame(self.root, text="Template Mapping")
        map_frame.pack(fill="x", padx=10, pady=8)

        ttk.Label(map_frame, text="Sheet name").grid(row=0, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(map_frame, textvariable=self.sheet_name, width=16).grid(row=0, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(map_frame, text="Header row").grid(row=0, column=2, sticky="w", padx=10, pady=6)
        ttk.Entry(map_frame, textvariable=self.header_row, width=10).grid(row=0, column=3, sticky="w", padx=10, pady=6)

        ttk.Label(map_frame, text="Start writing at row").grid(row=0, column=4, sticky="w", padx=10, pady=6)
        ttk.Entry(map_frame, textvariable=self.start_row, width=10).grid(row=0, column=5, sticky="w", padx=10, pady=6)

        ttk.Label(map_frame, text="Target column header").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(map_frame, textvariable=self.target_header, width=22).grid(row=1, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(map_frame, text="Default: Description").grid(row=1, column=2, columnspan=2, sticky="w", padx=10, pady=6)

        ttk.Checkbutton(
            map_frame,
            text="Prefix OCR text with image filename",
            variable=self.add_filename_prefix
        ).grid(row=2, column=0, columnspan=3, sticky="w", padx=10, pady=6)

        ttk.Checkbutton(
            map_frame,
            text="Insert filename into Item column",
            variable=self.insert_filename_item
        ).grid(row=2, column=3, columnspan=3, sticky="w", padx=10, pady=6)

        # OCR
        ocr_frame = ttk.LabelFrame(self.root, text="OCR Settings")
        ocr_frame.pack(fill="x", padx=10, pady=8)

        ttk.Label(ocr_frame, text="Tesseract path (optional)").grid(row=0, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(ocr_frame, textvariable=self.tesseract_path, width=68).grid(
            row=0, column=1, sticky="ew", padx=10, pady=6
        )
        ttk.Button(ocr_frame, text="Browse", command=self._browse_tesseract).grid(
            row=0, column=2, sticky="ew", padx=10, pady=6
        )

        ttk.Label(ocr_frame, text="Language").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        ttk.Entry(ocr_frame, textvariable=self.lang, width=12).grid(row=1, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(ocr_frame, text="PSM").place(in_=ocr_frame, relx=0.55, y=46)
        ttk.Entry(ocr_frame, textvariable=self.psm, width=8).place(in_=ocr_frame, relx=0.60, y=46)

        ttk.Label(ocr_frame, text="OEM").place(in_=ocr_frame, relx=0.72, y=46)
        ttk.Entry(ocr_frame, textvariable=self.oem, width=8).place(in_=ocr_frame, relx=0.77, y=46)

        ttk.Checkbutton(
            ocr_frame,
            text="Flatten line breaks into one paragraph",
            variable=self.clean_line_breaks
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=10, pady=6)

        ocr_frame.columnconfigure(1, weight=1)

        # Actions
        action_frame = ttk.Frame(self.root)
        action_frame.pack(fill="x", padx=10, pady=8)

        self.run_button = ttk.Button(action_frame, text="Create Filled Workbook", command=self._start_processing)
        self.run_button.pack(side="left", padx=(0, 8))

        ttk.Button(action_frame, text="Clear Log", command=self._clear_log).pack(side="left")

        # Progress
        progress_frame = ttk.LabelFrame(self.root, text="Progress")
        progress_frame.pack(fill="x", padx=10, pady=8)

        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=10)

        self.status_label = ttk.Label(progress_frame, text="Ready")
        self.status_label.pack(anchor="w", padx=10, pady=(0, 10))

        # Log
        log_frame = ttk.LabelFrame(self.root, text="Log")
        log_frame.pack(fill="both", expand=True, padx=10, pady=8)

        self.log_text = tk.Text(log_frame, wrap="word", height=18)
        self.log_text.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self._toggle_input_mode()

    def _log(self, text):
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.delete("1.0", "end")

    def _toggle_input_mode(self):
        mode = self.input_mode.get()
        if mode == "folder":
            self.folder_entry.configure(state="normal")
            self.folder_button.configure(state="normal")
            self.files_button.configure(state="disabled")
        else:
            self.folder_entry.configure(state="disabled")
            self.folder_button.configure(state="disabled")
            self.files_button.configure(state="normal")

    def _browse_template(self):
        path = filedialog.askopenfilename(
            title="Choose inventory Excel template",
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.template_file.set(path)
            if not self.output_file.get().strip():
                suggested = str(Path(path).with_name(Path(path).stem + "_filled.xlsx"))
                self.output_file.set(suggested)

    def _save_output(self):
        path = filedialog.asksaveasfilename(
            title="Save output workbook as",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")]
        )
        if path:
            self.output_file.set(path)

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Choose folder with JPG files")
        if folder:
            self.input_folder.set(folder)

    def _browse_files(self):
        files = filedialog.askopenfilenames(
            title="Choose JPG files",
            filetypes=[("JPEG files", "*.jpg *.jpeg"), ("All files", "*.*")]
        )
        if files:
            self.selected_files = list(files)
            self.files_label.config(text=f"{len(self.selected_files)} files selected")

    def _browse_tesseract(self):
        path = filedialog.askopenfilename(
            title="Select tesseract executable",
            filetypes=[("Executable", "*.exe"), ("All files", "*.*")]
        )
        if path:
            self.tesseract_path.set(path)

    def _validate(self):
        template = self.template_file.get().strip()
        if not template:
            raise ValueError("Choose the Excel template workbook.")
        if not Path(template).exists():
            raise ValueError("The Excel template file does not exist.")

        output = self.output_file.get().strip()
        if not output:
            raise ValueError("Choose where to save the output workbook.")
        if not output.lower().endswith(".xlsx"):
            raise ValueError("The output workbook must end with .xlsx")

        if self.input_mode.get() == "folder":
            folder = self.input_folder.get().strip()
            if not folder:
                raise ValueError("Choose a folder of JPG files.")
            if not Path(folder).exists():
                raise ValueError("The image folder does not exist.")
        else:
            if not self.selected_files:
                raise ValueError("Choose one or more JPG files.")


        if not self._ensure_tesseract():
            raise ValueError("Tesseract is not installed or it's not in your PATH. Please select the Tesseract executable.")

        try:
            int(self.header_row.get())
            int(self.start_row.get())
            int(self.psm.get())
            int(self.oem.get())
        except ValueError:
            raise ValueError("Header row, start row, PSM, and OEM must be integers.")

    def _get_image_files(self):
        if self.input_mode.get() == "folder":
            folder = Path(self.input_folder.get().strip())
            files = sorted(list(folder.glob("*.jpg")) + list(folder.glob("*.jpeg")))
        else:
            files = [Path(f) for f in self.selected_files]
        return files

    def _preprocess_image(self, img_path):
        img = Image.open(img_path)
        img = ImageOps.exif_transpose(img)
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    def _normalize_text(self, text):
        text = text.strip()
        if self.clean_line_breaks.get():
            text = re.sub(r"\s*\n\s*", " ", text)
            text = re.sub(r"\s{2,}", " ", text).strip()
        return text

    def _ocr_text(self, img_path):
        img = self._preprocess_image(img_path)
        config = f"--oem {self.oem.get()} --psm {self.psm.get()}"
        text = pytesseract.image_to_string(img, lang=self.lang.get().strip(), config=config)
        text = self._normalize_text(text)
        if self.add_filename_prefix.get():
            text = f"{img_path.name}: {text}" if text else img_path.name
        return text

    def _find_header_column(self, ws, header_row, header_text):
        target = str(header_text).strip().lower()
        for cell in ws[header_row]:
            value = "" if cell.value is None else str(cell.value).strip().lower()
            if value == target:
                return cell.column
        return None

    def _copy_row_style(self, ws, source_row, target_row):
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col)
            dst = ws.cell(row=target_row, column=col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)

    def _start_processing(self):
        if self.is_running:
            return
        try:
            self._validate()
        except Exception as e:
            messagebox.showerror("Validation Error", str(e))
            return

        self.is_running = True
        self.run_button.configure(state="disabled")
        threading.Thread(target=self._process, daemon=True).start()

    def _process(self):
        try:
            files = self._get_image_files()
            if not files:
                raise ValueError("No JPG or JPEG files were found.")

            template = Path(self.template_file.get().strip())
            output = Path(self.output_file.get().strip())
            sheet_name = self.sheet_name.get().strip()
            header_row = int(self.header_row.get())
            start_row = int(self.start_row.get())
            target_header = self.target_header.get().strip()

            wb = load_workbook(template)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' was not found in the workbook.")

            ws = wb[sheet_name]
            target_col = self._find_header_column(ws, header_row, target_header)
            if target_col is None:
                raise ValueError(f"Column header '{target_header}' was not found on row {header_row} in sheet '{sheet_name}'.")

            # Find the 'Item' column
            item_col = self._find_header_column(ws, header_row, 'item')
            if item_col is None:
                raise ValueError(f"Column header 'Item' was not found on row {header_row} in sheet '{sheet_name}'.")

            self.progress["maximum"] = len(files)
            self.progress["value"] = 0
            self._set_status(f"Processing {len(files)} images...")

            style_row = start_row if start_row <= ws.max_row else header_row + 1
            if style_row < 1:
                style_row = 1

            for idx, img_path in enumerate(files, start=1):
                row_num = start_row + idx - 1
                if row_num > ws.max_row:
                    ws.insert_rows(row_num, 1)
                    self._copy_row_style(ws, style_row, row_num)

                self._log(f"OCR: {img_path.name}")
                try:
                    text = self._ocr_text(img_path)
                except Exception as e:
                    text = f"ERROR: {e}"


                # Write OCR result to Description (target_col)
                cell = ws.cell(row=row_num, column=target_col)
                cell.value = text
                cell.alignment = copy(cell.alignment) if cell.alignment else copy(ws.cell(row=style_row, column=target_col).alignment)
                if cell.alignment:
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")

                # Optionally write filename to Item column
                if self.insert_filename_item.get():
                    item_cell = ws.cell(row=row_num, column=item_col)
                    item_cell.value = img_path.name
                    item_cell.alignment = copy(item_cell.alignment) if item_cell.alignment else copy(ws.cell(row=style_row, column=item_col).alignment)
                    if item_cell.alignment:
                        item_cell.alignment = copy(item_cell.alignment)
                        item_cell.alignment = item_cell.alignment.copy(wrapText=True, vertical="top")

                self.progress["value"] = idx
                self._set_status(f"Wrote row {row_num} ({idx} of {len(files)})")

            wb.save(output)
            self._log("")
            self._log(f"Done. Saved workbook: {output}")
            self._set_status("Completed successfully")
            self.root.after(0, lambda: messagebox.showinfo("Finished", f"Completed.\n\nSaved workbook:\n{output}"))

        except Exception as e:
            self._log(f"ERROR: {e}")
            self._set_status("Failed")
            self.root.after(0, lambda e=e: messagebox.showerror("Error", str(e)))
        finally:
            self.is_running = False
            self.root.after(0, lambda: self.run_button.configure(state="normal"))

    def _set_status(self, text):
        self.root.after(0, lambda: self.status_label.config(text=text))


def main():
    root = tk.Tk()
    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass
    app = InventoryOCRApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
